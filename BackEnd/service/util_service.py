from io                 import BytesIO
from datetime           import date
from uuid               import uuid1
from flask              import send_file
from flask_caching      import Cache

from openpyxl              import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from model.master_dao import MasterDao

cache = Cache(config={"CACHE_TYPE" : "simple"})

class UtilService:
    def excel_download(self, titles, data):
        """엑셀 파일을 만들고 전송하는 service
        Author:
            서득영

        Args:
            titles : 첫 행에 들어갈 컬럼 제목 리스트 ['a', 'b', 'c']
            data : 각 셀에 입력될 데이터 [{}, {}, {}]

        Returns:
            생성된 엑셀 파일 전송
        """
        workbook  = Workbook()
        worksheet = workbook.active

        today = str(date.today())
        uuid  = str(uuid1())
        
        file_name = today + "_" + uuid

        row = 1
        column = 1

        for title in titles:
            worksheet.cell(row, column, title)
            column += 1
        
        row = 2
        
        for items in data:
            column = 1
            for item in items.values():
                worksheet.cell(row, column, item)
                column += 1
            row += 1
        
        excel_file = BytesIO(save_virtual_workbook(workbook))

        return send_file(
                        excel_file,
                        attachment_filename=f"{file_name}.xlsx",
                        as_attachment=True
                        )
    
    @cache.cached(timeout=5*60)
    def get_action_dict(self, connection):
        """셀러 상태에 따른 마스터 액션 목록 생성

        Author:
            김현영

        Args:
            connection (객체): pymysql 객체 

        Returns:
            dict: {
                셀러 상태 아이디 : [{"action_id": 마스터 액션 번호, "action": 마스터 액션}]
                }
        """
        master_dao = MasterDao()
        actions    = master_dao.get_master_action_list(connection)
        
        actions_dict = {}
        for action in actions:
            if action["action_status_id"] not in actions_dict:
                actions_dict[action["action_status_id"]] = []        
            actions_dict[action["action_status_id"]].append({"action_id" : action["master_action_id"],
                                                                    "action" : action["action"]})

        return actions_dict
